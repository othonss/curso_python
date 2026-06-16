from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet

ROOT_FOLDER = Path(__file__).parent
WORKBOOK_PATH = ROOT_FOLDER / 'workbook.xlsx'

# Carregando um arquivo do Excel
workbook: Workbook = load_workbook(WORKBOOK_PATH)

#Nome para a planilha
sheet_name = 'Minha planilha'

#Selecionando a planilha
worksheet: Worksheet = workbook[sheet_name]

row: tuple[Cell]
for row in worksheet.iter_rows(min_row=2):
    for cell in row:
        print(cell.value, end='\t')

        #Alterando valor com base na condição
        if cell.value == 'Alberto':
            worksheet.cell(cell.row, 2, 30)
    print()

#Alterando valor na planilha
worksheet['B3'].value = 14
workbook.save(WORKBOOK_PATH)